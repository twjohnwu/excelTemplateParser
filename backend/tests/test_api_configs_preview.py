"""Tests for POST /api/configs/preview — stateless dry-run of the pipeline.

The preview endpoint must produce the same mapped rows the worker produces,
without creating a job, writing under /data, or touching Redis. The consistency
test pins that invariant against the real worker `_execute`.
"""

from __future__ import annotations

import json

from openpyxl import load_workbook

from app.schemas import ConfigSchema

# Same shape as test_worker_pipeline._config so preview and worker are comparable.
CONFIG = {
    "name": "demo",
    "target_template": {
        "sheet": "Sheet1",
        "header_row": 1,
        "columns": ["訂單編號", "客戶名稱", "業務員", "金額"],
    },
    "sources": [
        {"alias": "orders", "role": "primary", "sheet": "訂單", "header_row": 1},
        {"alias": "customers", "role": "lookup", "sheet": "客戶", "header_row": 1},
        {"alias": "sales", "role": "lookup", "sheet": "業務員", "header_row": 1},
    ],
    "joins": [
        {"left": "orders.客戶代號", "right": "customers.代號", "type": "left"},
        {"left": "customers.業務員代號", "right": "sales.代號", "type": "left"},
    ],
    "mappings": [
        {"target": "訂單編號", "source": "orders.單號"},
        {
            "target": "客戶名稱", "source": "customers.名稱",
            "conditions": [{"field": "orders.狀態", "op": "!=", "value": "取消"}],
            "default": "",
        },
        {"target": "業務員", "source": "sales.姓名"},
        {
            "target": "金額", "source": "orders.總額",
            "conditions": [{"field": "orders.總額", "op": ">=", "value": 0}],
            "default": 0,
        },
    ],
}


def _preview_files(target, orders, customers, sales):
    return [
        ("target_template", ("target.xlsx", open(target, "rb"), "application/octet-stream")),
        ("sources[orders]", ("orders.xlsx", open(orders, "rb"), "application/octet-stream")),
        ("sources[customers]", ("customers.xlsx", open(customers, "rb"), "application/octet-stream")),
        ("sources[sales]", ("sales.xlsx", open(sales, "rb"), "application/octet-stream")),
    ]


def test_preview_happy_path(api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx):
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG), "n": 20},
        files=files,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["columns"] == ["訂單編號", "客戶名稱", "業務員", "金額"]
    assert body["truncated"] is False
    rows = body["rows"]
    assert len(rows) == 3
    assert rows[0] == ["A001", "客戶一", "張三", 1000]
    # A002 is 取消 → 客戶名稱 condition fails → default "".
    assert rows[1][0] == "A002"
    assert rows[1][1] == ""
    assert rows[2] == ["A003", "客戶一", "張三", 2500]


def test_preview_no_side_effects(api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx):
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG)},
        files=files,
    )
    assert r.status_code == 200
    # No job created, nothing under /data/jobs, nothing enqueued.
    jobs_dir = api_client.data_dir / "jobs"
    assert not jobs_dir.exists() or not any(jobs_dir.iterdir())
    assert api_client.recorded_queue.enqueued == []


def test_preview_n_cap_and_truncated(api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx):
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG), "n": 2},
        files=files,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["rows"]) == 2  # capped at n
    assert body["truncated"] is True  # a 3rd row exists


def test_preview_422_on_missing_column(
    api_client, target_xlsx, customers_xlsx, sales_xlsx, make_xlsx
):
    broken = make_xlsx({
        "訂單": [["單號", "狀態", "總額"], ["A001", "成立", 100]],
    }, filename="broken_orders.xlsx")
    files = _preview_files(target_xlsx, broken, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG)},
        files=files,
    )
    assert r.status_code == 422
    detail = r.json()["detail"]
    assert "客戶代號" in detail["error"]
    assert detail["code"] == "TemplateInvalid"


def test_preview_422_on_missing_sheet(
    api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx
):
    cfg = json.loads(json.dumps(CONFIG))
    cfg["sources"][0]["sheet"] = "不存在的表"
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(cfg)},
        files=files,
    )
    assert r.status_code == 422
    assert r.json()["detail"]["code"] == "TemplateInvalid"


def test_preview_422_on_bad_config_json(
    api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx
):
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": "{not valid json"},
        files=files,
    )
    assert r.status_code == 422
    assert "config_json" in r.json()["detail"]


def test_preview_missing_source_file_422(api_client, target_xlsx, orders_xlsx, customers_xlsx):
    # sales source declared in config but no file uploaded.
    files = [
        ("target_template", ("target.xlsx", open(target_xlsx, "rb"), "application/octet-stream")),
        ("sources[orders]", ("orders.xlsx", open(orders_xlsx, "rb"), "application/octet-stream")),
        ("sources[customers]", ("customers.xlsx", open(customers_xlsx, "rb"), "application/octet-stream")),
    ]
    r = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG)},
        files=files,
    )
    assert r.status_code == 422
    assert "sales" in r.json()["detail"]


def test_preview_matches_worker_pipeline(
    tmp_path, redis_client, monkeypatch,
    api_client, target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx,
):
    """The preview rows equal the first rows the real worker `_execute` writes
    for the same config + files — the anti-drift guarantee."""
    from app.services.config_service import ConfigService
    from app.services.job_service import JobService
    from app.services.cleanup_service import CleanupService
    from app.settings import Settings
    from app.workers import tasks as worker_tasks

    # --- Run the real worker pipeline to produce an output xlsx. ---
    settings = Settings(REDIS_URL="redis://localhost:0", DATA_DIR=str(tmp_path / "wdata"))
    cfg_svc = ConfigService(redis=redis_client, configs_dir=settings.configs_dir)
    job_svc = JobService(redis=redis_client, jobs_dir=settings.jobs_dir)
    cleanup = CleanupService(
        redis=redis_client, jobs_dir=settings.jobs_dir, grace_minutes=60, retention_hours=24
    )
    monkeypatch.setattr(
        worker_tasks, "_wire",
        lambda: (redis_client, cfg_svc, job_svc, cleanup, settings),
    )
    config = ConfigSchema.model_validate(CONFIG)
    job_dir = settings.jobs_dir / "job1"
    (job_dir / "uploads" / "primary").mkdir(parents=True, exist_ok=True)
    (job_dir / "uploads" / "lookup").mkdir(parents=True, exist_ok=True)
    (job_dir / "out").mkdir(parents=True, exist_ok=True)
    (job_dir / "uploads" / "target.xlsx").write_bytes(target_xlsx.read_bytes())
    (job_dir / "uploads" / "primary" / "orders.xlsx").write_bytes(orders_xlsx.read_bytes())
    (job_dir / "uploads" / "lookup" / "customers.xlsx").write_bytes(customers_xlsx.read_bytes())
    (job_dir / "uploads" / "lookup" / "sales.xlsx").write_bytes(sales_xlsx.read_bytes())

    out_path = job_dir / "out" / "orders.xlsx.out.xlsx"
    worker_tasks._execute(job_dir, "orders.xlsx", config, out_path, chunk_size=10000)

    wb = load_workbook(out_path, data_only=True)
    ws = wb["Sheet1"]
    worker_rows = []
    for r_i in range(2, ws.max_row + 1):
        worker_rows.append([ws.cell(r_i, c).value for c in range(1, 5)])

    # --- Run preview via the API. ---
    files = _preview_files(target_xlsx, orders_xlsx, customers_xlsx, sales_xlsx)
    resp = api_client.post(
        "/api/configs/preview",
        data={"config_json": json.dumps(CONFIG), "n": 100},
        files=files,
    )
    assert resp.status_code == 200, resp.text
    preview_rows = resp.json()["rows"]

    # Normalize xlsx blank cells (None) vs preview "" for the conditional column.
    def _norm(v):
        return "" if v is None else v

    assert len(preview_rows) == len(worker_rows)
    for p_row, w_row in zip(preview_rows, worker_rows):
        assert [_norm(x) for x in p_row] == [_norm(x) for x in w_row]
