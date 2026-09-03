"""Worker pipeline tests: run_subtask end-to-end + finalize_job packaging.

These tests don't run an RQ worker — they call the task functions directly
after `monkeypatch`-ing `_wire()` to return our fakeredis + tmp paths.
"""

from __future__ import annotations

import contextlib
import threading
import time
import zipfile

import pytest
from openpyxl import load_workbook

from app.schemas import ConfigSchema
from app.services.cleanup_service import CleanupService
from app.services.config_service import ConfigService
from app.services.job_service import JobService
from app.settings import Settings
from app.workers import tasks as worker_tasks


def _config(preserve_styles: bool = True) -> ConfigSchema:
    return ConfigSchema.model_validate({
        "name": "demo",
        "target_template": {
            "sheet": "Sheet1", "header_row": 1,
            "columns": ["訂單編號", "客戶名稱", "業務員", "金額"],
            "preserve_styles": preserve_styles,
        },
        "sources": [
            {"alias": "orders", "role": "primary", "sheet": "訂單", "header_row": 1},
            {"alias": "customers", "role": "lookup", "sheet": "客戶", "header_row": 1},
            {"alias": "sales", "role": "lookup", "sheet": "業務員", "header_row": 1},
        ],
        "joins": [
            {"left": "orders.客戶代號", "right": "customers.代號", "type": "left"},
            {"left": "customers.業務員代號", "right": "sales.代號", "type": "left"},
        ],
        "mappings": [
            {"target": "訂單編號", "source": "orders.單號"},
            {
                "target": "客戶名稱", "source": "customers.名稱",
                "conditions": [{"field": "orders.狀態", "op": "!=", "value": "取消"}],
                "default": "",
            },
            {"target": "業務員", "source": "sales.姓名"},
            {
                "target": "金額", "source": "orders.總額",
                "conditions": [{"field": "orders.總額", "op": ">=", "value": 0}],
                "default": 0,
            },
        ],
    })


@pytest.fixture
def wired(monkeypatch, tmp_path, redis_client):
    """Patch worker_tasks._wire so run_subtask/finalize use our test setup."""
    settings = Settings(
        REDIS_URL="redis://localhost:0",
        DATA_DIR=str(tmp_path / "data"),
    )
    cfg_svc = ConfigService(redis=redis_client, configs_dir=settings.configs_dir)
    job_svc = JobService(redis=redis_client, jobs_dir=settings.jobs_dir)
    cleanup = CleanupService(
        redis=redis_client, jobs_dir=settings.jobs_dir,
        grace_minutes=60, retention_hours=24,
    )
    monkeypatch.setattr(worker_tasks, "_wire", lambda: (redis_client, cfg_svc, job_svc, cleanup, settings))
    return {
        "settings": settings,
        "redis": redis_client,
        "config_svc": cfg_svc,
        "job_svc": job_svc,
        "cleanup": cleanup,
        "tmp_path": tmp_path,
    }


def _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx, primary_name="orders.xlsx"):
    config = _config()
    wired["config_svc"].save(config)
    wired["job_svc"].create("job1", config, [primary_name])

    job_dir = wired["settings"].jobs_dir / "job1"
    (job_dir / "uploads" / "primary").mkdir(parents=True, exist_ok=True)
    (job_dir / "uploads" / "lookup").mkdir(parents=True, exist_ok=True)
    (job_dir / "out").mkdir(parents=True, exist_ok=True)

    (job_dir / "uploads" / "target.xlsx").write_bytes(target_xlsx.read_bytes())
    (job_dir / "uploads" / "primary" / primary_name).write_bytes(orders_xlsx.read_bytes())
    (job_dir / "uploads" / "lookup" / "customers.xlsx").write_bytes(customers_xlsx.read_bytes())
    (job_dir / "uploads" / "lookup" / "sales.xlsx").write_bytes(sales_xlsx.read_bytes())
    return config


def test_run_subtask_produces_correct_output(
    wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)

    worker_tasks.run_subtask("job1", "orders.xlsx")

    out = wired["settings"].jobs_dir / "job1" / "out" / "orders.xlsx.out.xlsx"
    assert out.exists()

    wb = load_workbook(out, data_only=True)
    ws = wb["Sheet1"]
    # Header preserved.
    assert ws.cell(1, 1).value == "訂單編號"
    # A001 row (成立, 1000): all fields populated.
    assert ws.cell(2, 1).value == "A001"
    assert ws.cell(2, 2).value == "客戶一"
    assert ws.cell(2, 3).value == "張三"
    assert ws.cell(2, 4).value == 1000
    # A002 row (取消): condition fails → 客戶名稱 empty, 金額 still passes (-?? actually 500≥0)
    # but 客戶名稱 mapping has conditions, so should be ""
    assert ws.cell(3, 2).value in (None, "")

    state = wired["job_svc"].get_state("job1")
    assert state.status == "done"
    assert state.subtasks["orders.xlsx"].status == "done"
    assert state.subtasks["orders.xlsx"].duration_ms is not None


def test_run_subtask_idempotent_skip(
    wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    out = wired["settings"].jobs_dir / "job1" / "out" / "orders.xlsx.out.xlsx"
    out.write_bytes(b"placeholder")  # pretend already done

    worker_tasks.run_subtask("job1", "orders.xlsx")

    # Output was NOT overwritten.
    assert out.read_bytes() == b"placeholder"
    assert wired["job_svc"].get_state("job1").subtasks["orders.xlsx"].status == "done"


def test_run_subtask_respects_cancel_flag(
    wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    wired["job_svc"].mark_cancelled("job1")
    worker_tasks.run_subtask("job1", "orders.xlsx")
    out = wired["settings"].jobs_dir / "job1" / "out" / "orders.xlsx.out.xlsx"
    assert not out.exists()  # cancelled before work


def test_run_subtask_records_failure(
    wired, customers_xlsx, sales_xlsx, target_xlsx, make_xlsx,
):
    # Primary missing 客戶代號 column triggers MappingError or similar inside join.
    broken = make_xlsx({
        "訂單": [["單號", "狀態", "總額"], ["A001", "成立", 100]],
    })
    _setup_job(wired, broken, customers_xlsx, sales_xlsx, target_xlsx, primary_name="broken.xlsx")

    with pytest.raises(Exception):
        worker_tasks.run_subtask("job1", "broken.xlsx")

    state = wired["job_svc"].get_state("job1")
    sub = state.subtasks["broken.xlsx"]
    assert sub.status == "failed"
    assert sub.user_message  # populated


def test_run_subtask_retries_transient_state_read_failure(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    """A bind-mount race can raise PermissionError from get_state before the
    main try/except. Two transient failures then a success should still let
    the subtask complete normally."""
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)

    real_get_state = JobService.get_state
    calls = {"n": 0}

    def flaky_get_state(self, job_id):
        calls["n"] += 1
        if calls["n"] <= 2:
            raise PermissionError("[Errno 1] Operation not permitted: state.json")
        return real_get_state(self, job_id)

    monkeypatch.setattr(JobService, "get_state", flaky_get_state)

    worker_tasks.run_subtask("job1", "orders.xlsx")

    out = wired["settings"].jobs_dir / "job1" / "out" / "orders.xlsx.out.xlsx"
    assert out.exists()
    assert calls["n"] == 3


def test_run_subtask_reraises_persistent_state_read_failure_without_mark_failed(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    """If state.json stays unreadable, run_subtask must re-raise (so RQ marks
    the job failed and the periodic scan re-enqueues it) WITHOUT calling
    mark_failed — state is exactly what's unreadable, and the subtask can
    still succeed on a later re-run."""
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)

    def always_fails(self, job_id):
        raise PermissionError("[Errno 1] Operation not permitted: state.json")

    monkeypatch.setattr(JobService, "get_state", always_fails)

    mark_failed_calls = []
    monkeypatch.setattr(
        JobService, "mark_failed",
        lambda self, *args, **kwargs: mark_failed_calls.append((args, kwargs)),
    )

    with pytest.raises(PermissionError):
        worker_tasks.run_subtask("job1", "orders.xlsx")

    assert mark_failed_calls == []


def test_finalize_packs_zip(
    wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    worker_tasks.run_subtask("job1", "orders.xlsx")
    worker_tasks.finalize_job("job1")

    zip_path = wired["settings"].jobs_dir / "job1" / "result.zip"
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    assert "orders.xlsx.out.xlsx" in names
    assert "_summary.txt" in names


def test_finalize_skips_when_subtasks_pending(wired):
    cfg = _config()
    wired["config_svc"].save(cfg)
    wired["job_svc"].create("j2", cfg, ["a.xlsx", "b.xlsx"])
    wired["job_svc"].mark_done("j2", "a.xlsx", 100)
    # b not done yet → finalize should no-op
    worker_tasks.finalize_job("j2")
    assert not (wired["settings"].jobs_dir / "j2" / "result.zip").exists()


def _read_data_rows(out_path):
    """Return data cell values (rows 2..end, cols 1..4) as a list of tuples."""
    wb = load_workbook(out_path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append(tuple(ws.cell(r, c).value for c in range(1, 5)))
    return rows


@pytest.mark.parametrize("preserve_styles", [True, False])
@pytest.mark.parametrize("chunk_size", [1, 2, 100])
def test_execute_streaming_equivalent_across_chunk_sizes(
    wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx, chunk_size, preserve_styles,
):
    """Output values are identical whether the primary streams in many small
    chunks or one big chunk, and regardless of preserve_styles."""
    config = _config(preserve_styles=preserve_styles)
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    job_dir = wired["settings"].jobs_dir / "job1"
    out_path = job_dir / "out" / f"cs{chunk_size}_ps{int(preserve_styles)}.out.xlsx"

    worker_tasks._execute(job_dir, "orders.xlsx", config, out_path, chunk_size=chunk_size)

    rows = _read_data_rows(out_path)
    # A001 (成立) fully populated; A002 (取消) → 客戶名稱 blank; A003 (成立).
    assert rows[0] == ("A001", "客戶一", "張三", 1000)
    assert rows[1][0] == "A002"
    assert rows[1][1] in (None, "")
    assert rows[2] == ("A003", "客戶一", "張三", 2500)
    assert len(rows) == 3


def test_finalize_job_called_twice_packs_once(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    worker_tasks.run_subtask("job1", "orders.xlsx")

    calls = []
    real_pack = worker_tasks.zipper.pack

    def spy_pack(*args, **kwargs):
        calls.append((args, kwargs))
        return real_pack(*args, **kwargs)

    monkeypatch.setattr(worker_tasks.zipper, "pack", spy_pack)

    worker_tasks.finalize_job("job1")
    worker_tasks.finalize_job("job1")

    assert len(calls) == 1
    zip_path = wired["settings"].jobs_dir / "job1" / "result.zip"
    assert zip_path.exists()


def test_finalize_job_skips_pack_when_zip_already_exists(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    worker_tasks.run_subtask("job1", "orders.xlsx")

    zip_path = wired["settings"].jobs_dir / "job1" / "result.zip"
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    zip_path.write_bytes(b"pre-existing zip")

    calls = []
    monkeypatch.setattr(
        worker_tasks.zipper, "pack",
        lambda *a, **k: calls.append((a, k)),
    )

    worker_tasks.finalize_job("job1")

    assert calls == []
    assert zip_path.read_bytes() == b"pre-existing zip"


def test_finalize_job_concurrent_calls_pack_exactly_once(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    """Two finalize_job calls racing (RQ can schedule the job-creation safety
    net and the last-subtask enqueue close together) must pack the zip only
    once — guarded by JobService.finalize_lock."""
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    worker_tasks.run_subtask("job1", "orders.xlsx")

    real_pack = worker_tasks.zipper.pack
    calls = []

    def slow_pack(*args, **kwargs):
        calls.append((args, kwargs))
        time.sleep(0.3)
        return real_pack(*args, **kwargs)

    monkeypatch.setattr(worker_tasks.zipper, "pack", slow_pack)

    threads = [threading.Thread(target=worker_tasks.finalize_job, args=("job1",)) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert len(calls) == 1
    zip_path = wired["settings"].jobs_dir / "job1" / "result.zip"
    assert zip_path.exists()


def test_finalize_job_concurrent_calls_pack_twice_without_lock(
    monkeypatch, wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx,
):
    """Sanity check for the test above: with `finalize_lock` stubbed to a
    no-op, the race is real and both threads pack — proving the assertion
    in the previous test would fail without the lock."""
    _setup_job(wired, orders_xlsx, customers_xlsx, sales_xlsx, target_xlsx)
    worker_tasks.run_subtask("job1", "orders.xlsx")

    real_pack = worker_tasks.zipper.pack
    calls = []

    def slow_pack(*args, **kwargs):
        calls.append((args, kwargs))
        time.sleep(0.3)
        return real_pack(*args, **kwargs)

    monkeypatch.setattr(worker_tasks.zipper, "pack", slow_pack)
    monkeypatch.setattr(
        wired["job_svc"], "finalize_lock", lambda job_id: contextlib.nullcontext()
    )

    threads = [threading.Thread(target=worker_tasks.finalize_job, args=("job1",)) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert len(calls) == 2


def test_finalize_skips_cancelled(wired):
    cfg = _config()
    wired["config_svc"].save(cfg)
    wired["job_svc"].create("j3", cfg, ["a.xlsx"])
    wired["job_svc"].mark_cancelled("j3")
    worker_tasks.finalize_job("j3")
    assert not (wired["settings"].jobs_dir / "j3" / "result.zip").exists()
