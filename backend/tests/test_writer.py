import pandas as pd
from openpyxl import load_workbook

from app.core import writer


def test_write_preserves_template_headers_and_fills_data(target_xlsx, tmp_path):
    df = pd.DataFrame({
        "訂單編號": ["A001", "A002"],
        "客戶名稱": ["客戶一", "客戶二"],
        "業務員": ["張三", "李四"],
        "金額": [1000, 2500],
    })

    out_path = tmp_path / "out.xlsx"
    writer.write(target_xlsx, df, sheet="Sheet1", header_row=1, out_path=out_path)

    wb = load_workbook(out_path, data_only=True)
    ws = wb["Sheet1"]
    # Header row unchanged
    assert ws.cell(row=1, column=1).value == "訂單編號"
    assert ws.cell(row=1, column=4).value == "金額"
    # Data starts at row 2
    assert ws.cell(row=2, column=1).value == "A001"
    assert ws.cell(row=2, column=4).value == 1000
    assert ws.cell(row=3, column=3).value == "李四"


def test_write_preserves_column_width(make_xlsx, tmp_path):
    src = make_xlsx({
        "Sheet1": [
            ["欄A", "欄B"],
        ]
    })
    # Inject a custom column width.
    from openpyxl import load_workbook as _lw
    wb = _lw(src)
    wb["Sheet1"].column_dimensions["A"].width = 42
    wb.save(src)
    wb.close()

    df = pd.DataFrame({"欄A": [1], "欄B": [2]})
    out_path = tmp_path / "out.xlsx"
    writer.write(src, df, sheet="Sheet1", header_row=1, out_path=out_path)

    out_wb = load_workbook(out_path)
    assert out_wb["Sheet1"].column_dimensions["A"].width == 42


def test_write_appends_unknown_columns(target_xlsx, tmp_path):
    """A mapping target not present in the template is appended as a new column
    at the right edge (header at header_row, values below)."""
    df = pd.DataFrame({
        "訂單編號": ["A001", "A002"],
        "新欄位": ["x", "y"],
    })
    out_path = tmp_path / "out.xlsx"
    writer.write(target_xlsx, df, sheet="Sheet1", header_row=1, out_path=out_path)

    wb = load_workbook(out_path)
    ws = wb["Sheet1"]
    # Existing column preserved.
    assert ws.cell(row=1, column=1).value == "訂單編號"
    assert ws.cell(row=2, column=1).value == "A001"
    # Template originally had 4 columns (訂單編號/客戶名稱/業務員/金額); the new
    # column should be appended at column 5.
    assert ws.cell(row=1, column=5).value == "新欄位"
    assert ws.cell(row=2, column=5).value == "x"
    assert ws.cell(row=3, column=5).value == "y"


def test_write_handles_nan_as_none(target_xlsx, tmp_path):
    import numpy as np
    df = pd.DataFrame({
        "訂單編號": ["A001"],
        "金額": [np.nan],
    })
    out_path = tmp_path / "out.xlsx"
    writer.write(target_xlsx, df, sheet="Sheet1", header_row=1, out_path=out_path)
    wb = load_workbook(out_path)
    assert wb["Sheet1"].cell(row=2, column=4).value is None


# ---------- write_stream (chunked) ----------


def test_write_stream_concatenates_chunks(target_xlsx, tmp_path):
    chunks = [
        pd.DataFrame({"訂單編號": ["A001", "A002"], "金額": [10, 20]}),
        pd.DataFrame({"訂單編號": ["A003"], "金額": [30]}),
    ]
    out_path = tmp_path / "out.xlsx"
    writer.write_stream(target_xlsx, chunks, sheet="Sheet1", header_row=1, out_path=out_path)

    ws = load_workbook(out_path, data_only=True)["Sheet1"]
    # Rows continuous across chunk boundary (offset must not reset per chunk).
    assert ws.cell(row=2, column=1).value == "A001"
    assert ws.cell(row=3, column=1).value == "A002"
    assert ws.cell(row=4, column=1).value == "A003"
    assert ws.cell(row=4, column=4).value == 30


def test_write_stream_fast_mode_values_and_header(target_xlsx, tmp_path):
    chunks = [pd.DataFrame({"訂單編號": ["A001", "A002"], "金額": [1000, 2500]})]
    out_path = tmp_path / "out.xlsx"
    writer.write_stream(
        target_xlsx, chunks, sheet="Sheet1", header_row=1, out_path=out_path, preserve_styles=False
    )
    ws = load_workbook(out_path, data_only=True)["Sheet1"]
    assert ws.cell(row=1, column=1).value == "訂單編號"   # template header copied
    assert ws.cell(row=1, column=4).value == "金額"
    assert ws.cell(row=2, column=1).value == "A001"
    assert ws.cell(row=2, column=4).value == 1000
    assert ws.cell(row=3, column=1).value == "A002"


def test_write_stream_fast_mode_preserves_column_width(make_xlsx, tmp_path):
    src = make_xlsx({"Sheet1": [["欄A", "欄B"]]})
    from openpyxl import load_workbook as _lw
    wb = _lw(src)
    wb["Sheet1"].column_dimensions["A"].width = 42
    wb.save(src)
    wb.close()

    out_path = tmp_path / "out.xlsx"
    writer.write_stream(
        src, [pd.DataFrame({"欄A": [1], "欄B": [2]})],
        sheet="Sheet1", header_row=1, out_path=out_path, preserve_styles=False,
    )
    assert load_workbook(out_path)["Sheet1"].column_dimensions["A"].width == 42


def test_write_stream_fast_mode_appends_unknown_columns(target_xlsx, tmp_path):
    chunks = [pd.DataFrame({"訂單編號": ["A001"], "新欄位": ["x"]})]
    out_path = tmp_path / "out.xlsx"
    writer.write_stream(
        target_xlsx, chunks, sheet="Sheet1", header_row=1, out_path=out_path, preserve_styles=False
    )
    ws = load_workbook(out_path)["Sheet1"]
    # Template has 4 columns; the new column lands at column 5.
    assert ws.cell(row=1, column=5).value == "新欄位"
    assert ws.cell(row=2, column=5).value == "x"
