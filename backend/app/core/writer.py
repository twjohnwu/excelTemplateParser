"""Write mapped rows to a target template, streaming chunk-by-chunk.

`write_stream` consumes an iterable of DataFrames (all sharing the same columns)
and writes them under the template's header, in one of two modes:

- preserve_styles=True (default): open the template with openpyxl's normal mode
  and fill rows, keeping column widths, fonts, merged cells and formulas intact.
  Faithful, but the whole output workbook stays in memory (not bounded).
- preserve_styles=False: build the output with openpyxl write_only mode, copying
  only the template's head rows (values) and column widths, then append data rows
  as they stream. Memory stays bounded; data-region per-cell styling, formulas and
  merged cells are NOT preserved.

`write` is a thin wrapper for the single-DataFrame, style-preserving case.
"""

from __future__ import annotations

from collections.abc import Iterable, Iterator
from pathlib import Path

import pandas as pd
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .exceptions import WriterError


def write(target_template: str | Path, df: pd.DataFrame, sheet: str, header_row: int, out_path: str | Path) -> None:
    """Style-preserving write of a single DataFrame (back-compat convenience)."""
    write_stream(target_template, [df], sheet=sheet, header_row=header_row, out_path=out_path, preserve_styles=True)


def write_stream(
    target_template: str | Path,
    chunks: Iterable[pd.DataFrame],
    sheet: str,
    header_row: int,
    out_path: str | Path,
    *,
    preserve_styles: bool = True,
) -> None:
    """Open `target_template`, fill rows starting at `header_row + 1` from `chunks`,
    save to `out_path`. All chunks must share the same columns; the first chunk
    decides the output column layout."""
    src = Path(target_template)
    dst = Path(out_path)

    it = iter(chunks)
    try:
        first = next(it)
    except StopIteration:
        first = None
    # Note: pulling `first` may run upstream parse/join/map (lazy generator) and
    # surface a CoreError here, before any workbook is opened.

    columns = list(first.columns) if first is not None else []
    data_chunks = _chain_first(first, it)

    if preserve_styles:
        _write_styled(src, dst, columns, data_chunks, sheet, header_row)
    else:
        _write_fast(src, dst, columns, data_chunks, sheet, header_row)


def _chain_first(first: pd.DataFrame | None, rest: Iterator[pd.DataFrame]) -> Iterator[pd.DataFrame]:
    if first is not None:
        yield first
    yield from rest


def _open_template(src: Path):
    # Normal mode (not read_only): the template holds only header/style rows, never
    # the output data, so this stays cheap — and read_only lacks column_dimensions.
    try:
        return load_workbook(filename=str(src), data_only=False, keep_vba=False)
    except (InvalidFileException, FileNotFoundError, OSError, BadZipFile) as exc:
        raise WriterError(
            user_message="目標範本檔損毀或無法開啟",
            tech_detail=f"{type(exc).__name__}: {exc}",
            path=str(src),
        ) from exc


def _write_styled(
    src: Path, dst: Path, columns: list[str], chunks: Iterator[pd.DataFrame], sheet: str, header_row: int
) -> None:
    wb = _open_template(src)
    try:
        if sheet not in wb.sheetnames:
            raise WriterError(
                user_message=f"目標範本不含工作表「{sheet}」",
                tech_detail=f"available: {wb.sheetnames}",
                sheet=sheet,
            )
        ws = wb[sheet]

        # Map target column names to their column index in the template.
        header_index = {
            _str(ws.cell(row=header_row, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }
        # Mappings whose target isn't a template header are appended as new columns.
        next_col = ws.max_column + 1
        for col_name in columns:
            if col_name not in header_index:
                ws.cell(row=header_row, column=next_col, value=col_name)
                header_index[col_name] = next_col
                next_col += 1

        row_offset = header_row + 1
        for chunk in chunks:
            for _, row in chunk.iterrows():
                for col_name in columns:
                    value = row[col_name]
                    if pd.isna(value):
                        value = None
                    ws.cell(row=row_offset, column=header_index[col_name], value=value)
                row_offset += 1

        dst.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dst))
    finally:
        wb.close()


def _write_fast(
    src: Path, dst: Path, columns: list[str], chunks: Iterator[pd.DataFrame], sheet: str, header_row: int
) -> None:
    head_rows, template_headers, col_widths = _read_template_head(src, sheet, header_row)

    name_to_pos: dict[str, int] = {}
    for pos, name in enumerate(template_headers):
        if name and name not in name_to_pos:
            name_to_pos[name] = pos
    new_cols = [c for c in columns if c not in name_to_pos]
    tmpl_width = len(template_headers)
    new_pos = {name: tmpl_width + k for k, name in enumerate(new_cols)}
    out_width = tmpl_width + len(new_cols)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet)
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # Head rows (1..header_row): values only. The header row is extended with the
    # names of any appended (non-template) columns. Rows above keep their values.
    for i in range(1, header_row + 1):
        src_row = head_rows[i - 1] if i - 1 < len(head_rows) else []
        out_row: list[object] = [None] * out_width
        for j, val in enumerate(src_row[:tmpl_width]):
            out_row[j] = val
        if i == header_row:
            for k, name in enumerate(new_cols):
                out_row[tmpl_width + k] = name
        ws.append(out_row)

    for chunk in chunks:
        for _, row in chunk.iterrows():
            out_row = [None] * out_width
            for col_name in columns:
                value = row[col_name]
                if pd.isna(value):
                    value = None
                pos = name_to_pos.get(col_name)
                if pos is None:
                    pos = new_pos[col_name]
                out_row[pos] = value
            ws.append(out_row)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))


def _read_template_head(
    src: Path, sheet: str, header_row: int
) -> tuple[list[list[object]], list[str], dict[str, float]]:
    """Read rows 1..header_row (values) and column widths from the template."""
    wb = _open_template(src)
    try:
        if sheet not in wb.sheetnames:
            raise WriterError(
                user_message=f"目標範本不含工作表「{sheet}」",
                tech_detail=f"available: {wb.sheetnames}",
                sheet=sheet,
            )
        ws = wb[sheet]
        head_rows: list[list[object]] = []
        template_headers: list[str] = []
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=header_row, values_only=True), start=1
        ):
            head_rows.append(list(row))
            if i == header_row:
                template_headers = [_str(cell) for cell in row]
        col_widths = {
            letter: dim.width
            for letter, dim in ws.column_dimensions.items()
            if dim.width is not None
        }
        return head_rows, template_headers, col_widths
    finally:
        wb.close()


def _str(value: object) -> str:
    return "" if value is None else str(value).strip()
