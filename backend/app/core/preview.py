"""Stateless dry-run preview of the parse → join → map pipeline.

Shares the *exact* join+map+source_cell logic the worker runs (see
`app/workers/tasks.py::_execute`). `decisions_log.md` records a past regression
where copy-pasted pipeline logic drifted from the worker; to avoid that, the
source-cell resolver lives here and the worker imports it, and the join+map
closure below mirrors `_execute` step for step.

The only intentional difference from the worker: we consume just the FIRST
chunk of the primary (capped at `n` rows) and never call `writer.write_stream`.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from . import joiner, mapper, parser
from .exceptions import MappingError
from ..schemas import ConfigSchema


def resolve_source_cells(
    config: ConfigSchema,
    source_files: dict[str, Path],
    source_sheets: dict[str, str],
) -> dict[int, object]:
    """Resolve every mapping.source_cell to its xlsx cell value, keyed by
    mapping index. Workbooks are cached so each source file is opened once.

    Shared by the worker (`tasks._execute`) and the preview endpoint so the
    two code paths resolve cell references identically.
    """
    needed = [
        (i, m.source_cell)
        for i, m in enumerate(config.mappings)
        if m.source_cell is not None
    ]
    if not needed:
        return {}

    cache: dict[Path, "openpyxl.Workbook"] = {}
    resolved: dict[int, object] = {}
    try:
        for idx, sc in needed:
            path = source_files[sc.alias]
            sheet_name = source_sheets[sc.alias]
            if path not in cache:
                cache[path] = openpyxl.load_workbook(
                    path, data_only=True, read_only=True
                )
            wb = cache[path]
            if sheet_name not in wb.sheetnames:
                raise MappingError(
                    user_message=f"來源「{sc.alias}」中找不到工作表「{sheet_name}」",
                    tech_detail=f"sheets={wb.sheetnames}",
                    mapping=config.mappings[idx].model_dump(),
                )
            try:
                value = wb[sheet_name][sc.address].value
            except (ValueError, KeyError) as exc:
                raise MappingError(
                    user_message=f"來源「{sc.alias}」的儲存格『{sc.address}』讀取失敗",
                    tech_detail=str(exc),
                    mapping=config.mappings[idx].model_dump(),
                ) from exc
            resolved[idx] = value
    finally:
        for wb in cache.values():
            wb.close()
    return resolved


def df_needed_aliases(config: ConfigSchema) -> set[str]:
    """Aliases that must be parsed into a DataFrame for the join/map pipeline.

    `source_cell`-only sources are excluded — their values are read directly via
    openpyxl in `resolve_source_cells` and never enter the joined DataFrame.
    """
    needed = {config.primary_alias}
    for j in config.joins:
        needed.add(j.left.split(".", 1)[0])
        needed.add(j.right.split(".", 1)[0])
    for m in config.mappings:
        if m.source:
            needed.add(m.source.split(".", 1)[0])
        for c in m.conditions:
            needed.add(c.field.split(".", 1)[0])
    return needed


def _primary_is_join_base(config: ConfigSchema) -> bool:
    """Streaming requires the primary to be the join root (joiner uses the first
    rule's left alias as the merge base). With no joins the primary is trivially
    the base. Otherwise the worker falls back to a full load to preserve
    semantics, and so must preview.
    """
    if not config.joins:
        return True
    return config.joins[0].left.split(".", 1)[0] == config.primary_alias


def preview_map(
    config: ConfigSchema,
    source_files: dict[str, Path],
    source_sheets: dict[str, str],
    n: int,
) -> tuple[list[str], list[list[Any]], bool]:
    """Run parse → join → map on (at most) the first `n` primary rows.

    Returns `(columns, rows, truncated)`:
      - `columns`: output column order (same as the worker's mapper output)
      - `rows`: up to `n` mapped rows, JSON-safe (NaN/NaT → None)
      - `truncated`: True if more primary rows exist beyond the returned window

    Mirrors `tasks._execute` exactly except it consumes only the first chunk
    (capped at `n`) and never writes an xlsx.
    """
    if n < 1:
        n = 1

    df_needed = df_needed_aliases(config)
    primary_spec = next(s for s in config.sources if s.role == "primary")
    primary_alias = config.primary_alias
    primary_path = source_files[primary_alias]

    # Load lookups whole (same as the worker), primary handled separately below.
    lookups: dict = {}
    for spec in config.sources:
        if spec.role != "primary" and spec.alias in df_needed:
            lookups[spec.alias] = parser.parse(
                source_files[spec.alias], spec.sheet, spec.header_row
            ).df

    pinned_cells = resolve_source_cells(config, source_files, source_sheets)
    joins = [j.model_dump() for j in config.joins]
    mappings = [m.model_dump() for m in config.mappings]

    def _map_chunk(primary_df):
        joined = joiner.join({**lookups, primary_alias: primary_df}, joins)
        return mapper.apply(
            joined, mappings, config.target_template.columns, pinned_cells=pinned_cells
        )

    if _primary_is_join_base(config):
        # Stream: pull one chunk sized at `n`, then peek for a second chunk to
        # decide `truncated` without loading the whole primary.
        chunks = parser.iter_chunks(
            primary_path, primary_spec.sheet, primary_spec.header_row, chunk_size=n
        )
        first = next(chunks)
        primary_df = first.df
        truncated = len(primary_df) > n
        if not truncated:
            # First chunk fit within n rows; a further non-empty chunk means more data.
            truncated = any(len(c.df) > 0 for c in chunks)
        primary_df = primary_df.iloc[:n]
    else:
        # Non-streaming fallback: the primary is NOT the join base, so join
        # semantics require the full primary DataFrame before mapping (matches
        # the worker's full-load branch in `_execute`). We load it all, map,
        # then head — correctness over memory here (preview is a small dry run).
        primary_df = parser.parse(
            primary_path, primary_spec.sheet, primary_spec.header_row
        ).df
        truncated = len(primary_df) > n
        primary_df = primary_df.iloc[:n]

    mapped = _map_chunk(primary_df)
    columns = list(mapped.columns)
    rows = [_json_safe_row(row) for row in mapped.itertuples(index=False, name=None)]
    return columns, rows, truncated


def _json_safe_row(row: tuple) -> list[Any]:
    """Convert a mapped row to JSON-safe primitives: NaN/NaT/pd.NA → None."""
    import pandas as pd

    out: list[Any] = []
    for v in row:
        try:
            if v is None or (pd.api.types.is_scalar(v) and pd.isna(v)):
                out.append(None)
                continue
        except (TypeError, ValueError):
            pass
        out.append(v)
    return out
