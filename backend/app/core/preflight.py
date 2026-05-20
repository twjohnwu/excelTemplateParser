"""Synchronous preflight: verify an uploaded xlsx matches a config's requirements.

Called by the API before enqueueing batch subtasks so bad uploads fail fast
with a precise 422 message rather than crashing the worker mid-job.
"""

from __future__ import annotations

from pathlib import Path

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .exceptions import TemplateInvalid


def preflight_check(
    file_path: str | Path,
    sheet: str,
    header_row: int,
    required_columns: list[str],
) -> None:
    """Raise TemplateInvalid if `file_path` cannot satisfy the given expectations."""
    if header_row < 1:
        raise TemplateInvalid(
            user_message="標頭列號必須 ≥ 1",
            tech_detail=f"header_row={header_row}",
            file=str(file_path),
        )

    try:
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except (InvalidFileException, FileNotFoundError, OSError, BadZipFile) as exc:
        raise TemplateInvalid(
            user_message=f"檔案『{Path(file_path).name}』損毀或非 xlsx 格式",
            tech_detail=f"{type(exc).__name__}: {exc}",
            file=str(file_path),
        ) from exc

    try:
        if sheet not in wb.sheetnames:
            raise TemplateInvalid(
                user_message=f"檔案『{Path(file_path).name}』缺少工作表「{sheet}」",
                tech_detail=f"available: {wb.sheetnames}",
                file=str(file_path),
                sheet=sheet,
            )

        ws = wb[sheet]
        headers: list[str] | None = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == header_row:
                headers = [_norm(cell) for cell in row]
                break

        if headers is None:
            raise TemplateInvalid(
                user_message=f"檔案『{Path(file_path).name}』第 {header_row} 列不存在",
                tech_detail="header_row beyond last row in sheet",
                file=str(file_path),
                sheet=sheet,
                header_row=header_row,
            )

        missing = [col for col in required_columns if col not in headers]
        if missing:
            raise TemplateInvalid(
                user_message=(
                    f"檔案『{Path(file_path).name}』的工作表「{sheet}」缺少欄位："
                    + "、".join(missing)
                ),
                tech_detail=f"present: {headers}",
                file=str(file_path),
                sheet=sheet,
                missing_columns=missing,
            )
    finally:
        wb.close()


def _norm(cell: object) -> str:
    return "" if cell is None else str(cell).strip()
