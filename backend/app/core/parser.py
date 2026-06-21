"""Excel parsing: workbook → DataFrame + header info.

`iter_chunks` streams a sheet in bounded-size batches via openpyxl read_only
mode, so a large primary file never fully materializes in memory. `parse` is a
convenience wrapper that concatenates those chunks into one DataFrame — fine for
small lookup tables, but the worker streams the primary through `iter_chunks`.

Raises TemplateInvalid for any structural problem (corrupt file, missing
sheet, header row out of range).
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .exceptions import TemplateInvalid

DEFAULT_CHUNK_SIZE = 10000


@dataclass(frozen=True)
class ParsedSheet:
    """Result of parsing one sheet of a workbook."""

    headers: list[str]
    df: pd.DataFrame


def list_sheets(path: str | Path) -> list[str]:
    """Return sheet names in the workbook (cheap; no row scan)."""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except (InvalidFileException, FileNotFoundError, OSError, KeyError, BadZipFile) as exc:
        raise TemplateInvalid(
            user_message="範本檔損毀或非 xlsx 格式",
            tech_detail=f"{type(exc).__name__}: {exc}",
            path=str(path),
        ) from exc
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def preview_rows(path: str | Path, sheet: str, *, start: int = 1, count: int = 30) -> list[list[object]]:
    """Return the first `count` rows of `sheet` starting from `start` (1-indexed).

    Used by the templates parse endpoint to power the sheet/header_row picker.
    """
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except (InvalidFileException, FileNotFoundError, OSError, BadZipFile) as exc:
        raise TemplateInvalid(
            user_message="範本檔損毀或非 xlsx 格式",
            tech_detail=f"{type(exc).__name__}: {exc}",
            path=str(path),
        ) from exc

    try:
        if sheet not in wb.sheetnames:
            raise TemplateInvalid(
                user_message=f"工作表「{sheet}」不存在",
                tech_detail=f"available sheets: {wb.sheetnames}",
                sheet=sheet,
            )
        ws = wb[sheet]
        out: list[list[object]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < start:
                continue
            out.append(list(row))
            if len(out) >= count:
                break
        return out
    finally:
        wb.close()


def iter_chunks(
    path: str | Path, sheet: str, header_row: int, *, chunk_size: int = DEFAULT_CHUNK_SIZE
) -> Iterator[ParsedSheet]:
    """Stream `sheet` from `path` in batches of at most `chunk_size` data rows.

    Row `header_row` (1-indexed) supplies the headers; rows above it are metadata
    and skipped; empty rows below it are dropped. Each yielded `ParsedSheet` shares
    the same `headers` and carries one batch as its `df` — the full row set is never
    held in memory at once. Always yields at least one chunk (an empty `df` when the
    sheet has headers but no data rows) so callers can rely on the column layout.

    NOTE: this is a generator — argument/structural validation happens lazily on the
    first `next()`, matching how the worker consumes it.
    """
    if chunk_size < 1:
        chunk_size = DEFAULT_CHUNK_SIZE
    if header_row < 1:
        raise TemplateInvalid(
            user_message="標頭列號必須 ≥ 1",
            tech_detail=f"header_row={header_row}",
        )

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except (InvalidFileException, FileNotFoundError, OSError, BadZipFile) as exc:
        raise TemplateInvalid(
            user_message="範本檔損毀或非 xlsx 格式",
            tech_detail=f"{type(exc).__name__}: {exc}",
            path=str(path),
        ) from exc

    try:
        if sheet not in wb.sheetnames:
            raise TemplateInvalid(
                user_message=f"工作表「{sheet}」不存在",
                tech_detail=f"available sheets: {wb.sheetnames}",
                sheet=sheet,
            )

        ws = wb[sheet]
        headers: list[str] | None = None
        buffer: list[list[object]] = []
        emitted = False
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < header_row:
                continue
            if i == header_row:
                headers = [_normalize_header(cell) for cell in row]
                continue
            if _row_is_empty(row):
                continue
            buffer.append(list(row))
            if len(buffer) >= chunk_size:
                yield _make_sheet(headers, buffer)
                emitted = True
                buffer = []

        if headers is None:
            raise TemplateInvalid(
                user_message=f"工作表「{sheet}」沒有第 {header_row} 列",
                tech_detail="header_row beyond last row in sheet",
                sheet=sheet,
                header_row=header_row,
            )

        if buffer:
            yield _make_sheet(headers, buffer)
            emitted = True
        if not emitted:
            yield _make_sheet(headers, [])
    finally:
        wb.close()


def parse(path: str | Path, sheet: str, header_row: int) -> ParsedSheet:
    """Read `sheet` from `path` as a single DataFrame, using row `header_row`
    (1-indexed) as headers.

    Convenience wrapper over `iter_chunks` for small sources (e.g. lookup tables);
    concatenates all chunks. For large inputs prefer `iter_chunks` to keep memory
    bounded. Rows above `header_row` are treated as metadata and skipped.
    """
    chunks = list(iter_chunks(path, sheet, header_row, chunk_size=DEFAULT_CHUNK_SIZE))
    if len(chunks) == 1:
        return chunks[0]
    headers = chunks[0].headers
    df = pd.concat([c.df for c in chunks], ignore_index=True)
    return ParsedSheet(headers=headers, df=df)


def _make_sheet(headers: list[str], data_rows: list[list[object]]) -> ParsedSheet:
    """Pad/trim each row to the header width and build a DataFrame for one chunk."""
    width = len(headers)
    normalized = [list(row[:width]) + [None] * max(0, width - len(row)) for row in data_rows]
    return ParsedSheet(headers=headers, df=pd.DataFrame(normalized, columns=headers))


def _normalize_header(cell: object) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _row_is_empty(row: tuple) -> bool:
    return all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row)
