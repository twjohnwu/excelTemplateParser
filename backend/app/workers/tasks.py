"""Worker pipeline: one subtask = one primary file → one output xlsx.

The whole pipeline lives here so it's the single assembly point referenced
by `decisions_log.md` (workers/tasks.py is the only place that strings
parser → joiner → mapper → writer together).

Idempotent: if `out/{source}.out.xlsx` already exists the subtask is a no-op
(the worker still calls mark_done so state.json reflects reality).
"""

from __future__ import annotations

import time
from pathlib import Path

import structlog
from redis import Redis

import openpyxl

from ..core import joiner, mapper, parser, writer, zipper
from ..core.exceptions import CoreError, MappingError
from ..schemas import ConfigSchema, JobState, SubtaskState
from ..services.cleanup_service import CleanupService
from ..services.config_service import ConfigService
from ..services.job_service import JobService
from ..settings import get_settings

log = structlog.get_logger(__name__)


def _wire():
    """Build service objects from settings. Workers create one per task call
    (cheap; Redis connection is reused via the connection pool)."""
    settings = get_settings()
    redis = Redis.from_url(settings.redis_url)
    return (
        redis,
        ConfigService(redis=redis, configs_dir=settings.configs_dir),
        JobService(redis=redis, jobs_dir=settings.jobs_dir),
        CleanupService(
            redis=redis,
            jobs_dir=settings.jobs_dir,
            grace_minutes=settings.download_grace_minutes,
            retention_hours=settings.job_retention_hours,
        ),
        settings,
    )


def run_subtask(job_id: str, source_file: str) -> None:
    """Process one primary file for a job.

    Steps: parse primary + all lookups → join → map → write.
    """
    redis, config_svc, job_svc, _cleanup, settings = _wire()
    bound = log.bind(job_id=job_id, source_file=source_file)

    # Honor cancellation BEFORE doing any work.
    if job_svc.is_cancel_requested(job_id):
        bound.info("subtask.skip", reason="cancelled")
        return

    job_dir = settings.jobs_dir / job_id
    out_path = job_dir / "out" / f"{source_file}.out.xlsx"
    if out_path.exists():
        bound.info("subtask.skip", reason="already_done")
        job_svc.mark_done(job_id, source_file, duration_ms=0)
        return

    state = job_svc.get_state(job_id)
    if state.config_name is None:
        job_svc.mark_failed(job_id, source_file, "Job 沒有對應的 config_name", tech_detail="state.config_name is None")
        raise RuntimeError("missing config_name")

    try:
        config = config_svc.get(state.config_name)
    except Exception as exc:
        job_svc.mark_failed(job_id, source_file, "找不到對應的設定檔", tech_detail=str(exc))
        raise

    started = time.monotonic()
    job_svc.mark_running(job_id, source_file)

    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        _execute(job_dir, source_file, config, out_path, chunk_size=settings.parse_chunk_size)
    except CoreError as exc:
        is_last = job_svc.mark_failed(job_id, source_file, exc.user_message, exc.tech_detail)
        bound.error("subtask.failed", **exc.context, exc_info=True)
        if is_last:
            from .queue import enqueue_finalize, make_queue
            enqueue_finalize(make_queue(redis), job_id)
        raise
    except Exception as exc:
        is_last = job_svc.mark_failed(job_id, source_file, "未預期錯誤，請聯絡管理員", tech_detail=repr(exc))
        bound.exception("subtask.unexpected")
        if is_last:
            from .queue import enqueue_finalize, make_queue
            enqueue_finalize(make_queue(redis), job_id)
        raise

    duration_ms = int((time.monotonic() - started) * 1000)
    is_last = job_svc.mark_done(job_id, source_file, duration_ms=duration_ms)
    bound.info("subtask.done", duration_ms=duration_ms, is_last=is_last)

    if is_last:
        # The job-creation path enqueues finalize_job up front as a safety net,
        # but that one races ahead of subtasks. Enqueue again from the last
        # completing subtask so finalize is guaranteed to run after all done.
        from .queue import enqueue_finalize, make_queue
        enqueue_finalize(make_queue(redis), job_id)


def df_needed_aliases(config: ConfigSchema) -> set[str]:
    """Aliases that need to be parsed into a DataFrame for the join/map pipeline.

    `source_cell`-only sources are excluded — their values are read directly via
    openpyxl in `_resolve_source_cells` and never enter the joined DataFrame.
    """
    needed = {config.primary_alias}
    for j in config.joins:
        needed.add(j.left.split(".", 1)[0])
        needed.add(j.right.split(".", 1)[0])
    for m in config.mappings:
        if m.source:
            needed.add(m.source.split(".", 1)[0])
        for c in m.conditions:
            needed.add(c.field.split(".", 1)[0])
    return needed


def _primary_is_join_base(config: ConfigSchema) -> bool:
    """Streaming requires the primary to be the join root (joiner uses the first
    rule's left alias as the merge base). With no joins the primary is trivially
    the base. Otherwise we must fall back to a full load to preserve semantics.
    """
    if not config.joins:
        return True
    return config.joins[0].left.split(".", 1)[0] == config.primary_alias


def _execute(
    job_dir: Path, primary_file: str, config: ConfigSchema, out_path: Path, *, chunk_size: int
) -> None:
    """Run parse → join → map → write for a single primary file.

    The primary is streamed in chunks (bounded memory); lookups are loaded whole
    and re-joined onto each chunk. Falls back to a full load when the primary
    isn't the join base (rare), so join semantics never change.
    """
    primary_path = job_dir / "uploads" / "primary" / primary_file
    target_path = job_dir / "uploads" / "target.xlsx"

    df_needed = df_needed_aliases(config)
    primary_spec = next(s for s in config.sources if s.role == "primary")

    lookups: dict = {}
    source_files: dict[str, Path] = {}
    source_sheets: dict[str, str] = {}
    for spec in config.sources:
        if spec.role == "primary":
            file_path = primary_path
        else:
            file_path = job_dir / "uploads" / "lookup" / f"{spec.alias}.xlsx"
        source_files[spec.alias] = file_path
        source_sheets[spec.alias] = spec.sheet
        if spec.role != "primary" and spec.alias in df_needed:
            lookups[spec.alias] = parser.parse(file_path, spec.sheet, spec.header_row).df

    pinned_cells = _resolve_source_cells(config, source_files, source_sheets)
    joins = [j.model_dump() for j in config.joins]
    mappings = [m.model_dump() for m in config.mappings]
    primary_alias = config.primary_alias

    def _map_chunk(primary_df):
        joined = joiner.join({**lookups, primary_alias: primary_df}, joins)
        return mapper.apply(joined, mappings, config.target_template.columns, pinned_cells=pinned_cells)

    if _primary_is_join_base(config):
        mapped_chunks = (
            _map_chunk(chunk.df)
            for chunk in parser.iter_chunks(
                primary_path, primary_spec.sheet, primary_spec.header_row, chunk_size=chunk_size
            )
        )
    else:
        primary_df = parser.parse(primary_path, primary_spec.sheet, primary_spec.header_row).df
        mapped_chunks = iter([_map_chunk(primary_df)])

    writer.write_stream(
        target_path,
        mapped_chunks,
        sheet=config.target_template.sheet,
        header_row=config.target_template.header_row,
        out_path=out_path,
        preserve_styles=config.target_template.preserve_styles,
    )


def _resolve_source_cells(
    config: ConfigSchema,
    source_files: dict[str, Path],
    source_sheets: dict[str, str],
) -> dict[int, object]:
    """Resolve every mapping.source_cell to its xlsx cell value, keyed by
    mapping index. Workbooks are cached so each source file is opened once.
    """
    needed = [
        (i, m.source_cell)
        for i, m in enumerate(config.mappings)
        if m.source_cell is not None
    ]
    if not needed:
        return {}

    cache: dict[Path, "openpyxl.Workbook"] = {}
    resolved: dict[int, object] = {}
    try:
        for idx, sc in needed:
            path = source_files[sc.alias]
            sheet_name = source_sheets[sc.alias]
            if path not in cache:
                cache[path] = openpyxl.load_workbook(
                    path, data_only=True, read_only=True
                )
            wb = cache[path]
            if sheet_name not in wb.sheetnames:
                raise MappingError(
                    user_message=f"來源「{sc.alias}」中找不到工作表「{sheet_name}」",
                    tech_detail=f"sheets={wb.sheetnames}",
                    mapping=config.mappings[idx].model_dump(),
                )
            try:
                value = wb[sheet_name][sc.address].value
            except (ValueError, KeyError) as exc:
                raise MappingError(
                    user_message=f"來源「{sc.alias}」的儲存格『{sc.address}』讀取失敗",
                    tech_detail=str(exc),
                    mapping=config.mappings[idx].model_dump(),
                ) from exc
            resolved[idx] = value
    finally:
        for wb in cache.values():
            wb.close()
    return resolved


def finalize_job(job_id: str) -> None:
    """Pack the per-subtask outputs into result.zip once every subtask is terminal.

    Idempotent: if already packed, returns early. Re-enqueues itself one
    iteration if some subtasks are still pending (RQ may schedule us early).
    """
    _redis, config_svc, job_svc, _cleanup, settings = _wire()
    state = job_svc.get_state(job_id)
    bound = log.bind(job_id=job_id)

    if state.status == "cancelled":
        bound.info("finalize.skip", reason="cancelled")
        return

    terminal = sum(1 for s in state.subtasks.values() if s.status in ("done", "failed"))
    if terminal < state.total:
        bound.info("finalize.early", terminal=terminal, total=state.total)
        return

    job_dir = settings.jobs_dir / job_id
    out_dir = job_dir / "out"
    zip_path = job_dir / "result.zip"

    if state.failed == state.total:
        bound.warning("finalize.all_failed", failed=state.failed)
        return

    summary = _build_summary(state, config_svc)
    zipper.pack(out_dir, zip_path, summary=summary)
    bound.info("finalize.packed", zip=str(zip_path))


def _build_summary(state: JobState, config_svc: ConfigService) -> str:
    lines = [
        f"Job: {state.job_id}",
        f"Config: {state.config_name}",
        f"Created: {state.created_at}",
        f"Total: {state.total}, Done: {state.done}, Failed: {state.failed}",
        "",
        "Subtasks:",
    ]
    for name, sub in state.subtasks.items():
        if sub.status == "done":
            lines.append(f"  ✓ {name}  ({sub.duration_ms} ms)")
        elif sub.status == "failed":
            lines.append(f"  ✗ {name}  — {sub.user_message}")
        else:
            lines.append(f"  ? {name}  status={sub.status}")
    return "\n".join(lines) + "\n"
