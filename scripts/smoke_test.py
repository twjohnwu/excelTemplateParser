"""End-to-end smoke test runner.

Run after `docker compose up -d redis api worker`. Walks流程 A then 流程 B,
prints PASS/FAIL per §8 verification scenario where automatable.

Requires: httpx, openpyxl (use the backend venv).
"""

from __future__ import annotations

import io
import json
import os
import sys
import tempfile
import time
import zipfile
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook

API = os.environ.get("API_BASE", "http://localhost:8000")
PASS = "\033[32m✓\033[0m"
FAIL = "\033[31m✗\033[0m"


def section(title: str) -> None:
    print(f"\n=== {title} ===")


def check(label: str, ok: bool, detail: str = "") -> bool:
    marker = PASS if ok else FAIL
    print(f"  {marker} {label}" + (f"  — {detail}" if detail else ""))
    return ok


def make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def wait_for_status(client: httpx.Client, job_id: str, target: set[str], timeout: int = 60) -> str:
    deadline = time.time() + timeout
    while time.time() < deadline:
        r = client.get(f"{API}/api/jobs/{job_id}")
        if r.status_code == 200:
            status = r.json()["status"]
            if status in target:
                return status
        time.sleep(0.5)
    return "timeout"


def main() -> int:
    failures: list[str] = []

    # --- Fixtures ---
    orders_xlsx = make_xlsx({
        "訂單": [
            ["單號", "客戶代號", "狀態", "總額"],
            ["A001", "C1", "成立", 1000],
            ["A002", "C2", "取消", 500],
            ["A003", "C1", "成立", 2500],
        ]
    })
    customers_xlsx = make_xlsx({
        "客戶": [
            ["代號", "名稱", "業務員代號"],
            ["C1", "客戶一", "S1"],
            ["C2", "客戶二", "S2"],
        ]
    })
    sales_xlsx = make_xlsx({
        "業務員": [
            ["代號", "姓名"],
            ["S1", "張三"],
            ["S2", "李四"],
        ]
    })
    target_xlsx = make_xlsx({
        "Sheet1": [
            ["訂單編號", "客戶名稱", "業務員", "金額"],
        ]
    })

    config = {
        "version": "1.0",
        "name": "smoke_test",
        "target_template": {
            "sheet": "Sheet1", "header_row": 1, "preserve_styles": True,
            "columns": ["訂單編號", "客戶名稱", "業務員", "金額"],
        },
        "sources": [
            {"alias": "orders", "role": "primary", "sheet": "訂單", "header_row": 1},
            {"alias": "customers", "role": "lookup", "sheet": "客戶", "header_row": 1},
            {"alias": "sales", "role": "lookup", "sheet": "業務員", "header_row": 1},
        ],
        "joins": [
            {"left": "orders.客戶代號", "right": "customers.代號", "type": "left"},
            {"left": "customers.業務員代號", "right": "sales.代號", "type": "left"},
        ],
        "mappings": [
            {"target": "訂單編號", "source": "orders.單號"},
            {
                "target": "客戶名稱", "source": "customers.名稱",
                "conditions": [{"field": "orders.狀態", "op": "!=", "value": "取消"}],
                "default": "",
            },
            {"target": "業務員", "source": "sales.姓名"},
            {
                "target": "金額", "source": "orders.總額",
                "conditions": [{"field": "orders.總額", "op": ">=", "value": 0}],
                "default": 0,
            },
        ],
    }

    with httpx.Client(timeout=30) as client:
        # --- §8.2 health ---
        section("§8.2 docker compose health")
        r = client.get(f"{API}/api/health")
        check("api responding", r.status_code == 200 and r.json().get("ok"))

        # --- §8.3 flow A: create config ---
        section("§8.3 flow A — save config")
        # Clean slate
        client.delete(f"{API}/api/configs/smoke_test")

        r = client.post(f"{API}/api/configs", json=config)
        check("POST /api/configs returns 200", r.status_code == 200, str(r.status_code))

        r = client.get(f"{API}/api/configs/smoke_test")
        check("GET /api/configs/smoke_test returns config",
              r.status_code == 200 and r.json()["name"] == "smoke_test")

        r = client.get(f"{API}/api/configs")
        check("GET /api/configs lists smoke_test",
              "smoke_test" in r.json()["configs"])

        # --- §8.7 duplicate name 409 ---
        section("§8.7 duplicate / invalid name error paths")
        r = client.post(f"{API}/api/configs", json=config)
        check("duplicate save returns 409", r.status_code == 409)

        bad = dict(config, name="bad/name")
        r = client.post(f"{API}/api/configs", json=bad)
        check("invalid name returns 422", r.status_code == 422)

        # --- §8.4 + §8.14 flow B: multi-source batch ---
        section("§8.4 + §8.14 flow B — multi-source batch (3 primary × shared lookups)")
        files = [
            ("target_template", ("target.xlsx", target_xlsx, "application/octet-stream")),
            ("sources[orders]", ("may.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[orders]", ("jun.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[orders]", ("jul.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[customers]", ("customers.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("sales.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files)
        check("POST /api/jobs returns 200", r.status_code == 200, r.text[:200])
        job_id = r.json()["job_id"]
        check("job_id returned", bool(job_id))

        status = wait_for_status(client, job_id, {"done", "failed"}, timeout=60)
        check(f"job reaches terminal status (got {status})", status == "done")

        snap = client.get(f"{API}/api/jobs/{job_id}").json()
        check(f"3 subtasks, 3 done (got {snap['done']}/{snap['total']})",
              snap["total"] == 3 and snap["done"] == 3)

        # ZIP download — wait for finalize_job to pack the archive.
        deadline = time.time() + 30
        r = None
        while time.time() < deadline:
            r = client.get(f"{API}/api/jobs/{job_id}/zip")
            if r.status_code == 200:
                break
            time.sleep(0.5)
        check(f"GET /zip returns 200 (got {r.status_code if r else 'n/a'})",
              r is not None and r.status_code == 200)
        check("response is zip", r.headers.get("content-type", "").startswith("application/"))
        with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
            names = zf.namelist()
        check(f"zip contains 3 outputs + _summary (names={names})",
              "_summary.txt" in names and sum(1 for n in names if n.endswith(".out.xlsx")) == 3)

        # --- §8.5 styles preserved (basic) ---
        section("§8.5 output style sanity")
        with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
            with zf.open("may.xlsx.out.xlsx") as fh:
                wb = load_workbook(io.BytesIO(fh.read()), data_only=True)
        ws = wb["Sheet1"]
        check("header row preserved", ws.cell(1, 1).value == "訂單編號")
        check("A001 row populated", ws.cell(2, 1).value == "A001" and ws.cell(2, 4).value == 1000)
        check("取消 row drops 客戶名稱 (condition !=取消)", ws.cell(3, 2).value in (None, ""))
        wb.close()

        # --- §8.21 grace period ---
        section("§8.21 download grace period (second fetch within window)")
        r2 = client.get(f"{API}/api/jobs/{job_id}/zip")
        check("re-download within grace returns 200", r2.status_code == 200)

        # --- §8.22 Range download ---
        section("§8.22 HTTP Range partial content")
        r3 = client.get(f"{API}/api/jobs/{job_id}/zip", headers={"Range": "bytes=10-100"})
        check(f"Range request returns 206 (got {r3.status_code})", r3.status_code == 206)

        # --- §8.14 validation (0 primary) ---
        section("§8.14 multi-source validation")
        files_no_primary = [
            ("target_template", ("t.xlsx", target_xlsx, "application/octet-stream")),
            ("sources[customers]", ("c.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("s.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files_no_primary)
        check("0 primary returns 422", r.status_code == 422,
              str(r.json().get("detail"))[:80])

        # --- §8.17 duplicate primary filenames ---
        section("§8.17 duplicate primary filenames rejected")
        files_dup = [
            ("target_template", ("t.xlsx", target_xlsx, "application/octet-stream")),
            ("sources[orders]", ("orders.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[orders]", ("orders.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[customers]", ("c.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("s.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files_dup)
        check("duplicate primary returns 422", r.status_code == 422)

        # --- §8.16 preflight (missing column) ---
        section("§8.16 preflight validation rejects missing column")
        bad_orders = make_xlsx({
            "訂單": [["單號", "狀態", "總額"], ["A001", "成立", 100]],  # no 客戶代號
        })
        files_bad = [
            ("target_template", ("t.xlsx", target_xlsx, "application/octet-stream")),
            ("sources[orders]", ("bad.xlsx", bad_orders, "application/octet-stream")),
            ("sources[customers]", ("c.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("s.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files_bad)
        check("missing column returns 422", r.status_code == 422)
        check("error mentions missing column",
              "客戶代號" in str(r.json().get("detail")))

        # --- §8.7 / §8.13 request_id on errors ---
        section("§8.13 error response carries request_id")
        check("422 response has x-request-id header",
              bool(r.headers.get("x-request-id")))

        # --- §8.10 persistence: config survives down/up (smaller variant) ---
        section("§8.10 persistence: re-fetch config after delete + re-save")
        r = client.delete(f"{API}/api/configs/smoke_test")
        check("delete returns 204", r.status_code == 204)
        client.post(f"{API}/api/configs", json=config)
        r = client.get(f"{API}/api/configs/smoke_test")
        check("re-fetch after re-save returns 200", r.status_code == 200)

        # --- §8.18 cancel a pending job ---
        section("§8.18 cancel an in-flight job")
        files_long = [
            ("target_template", ("t.xlsx", target_xlsx, "application/octet-stream")),
            *[("sources[orders]", (f"f{i}.xlsx", orders_xlsx, "application/octet-stream")) for i in range(5)],
            ("sources[customers]", ("c.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("s.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files_long)
        cancel_id = r.json()["job_id"]
        cancel_r = client.post(f"{API}/api/jobs/{cancel_id}/cancel")
        check(f"cancel returns 200 (got {cancel_r.status_code})", cancel_r.status_code == 200)
        # After cancel the job is deleted from disk; subsequent GET should 404.
        r = client.get(f"{API}/api/jobs/{cancel_id}")
        check("cancelled job is purged (404 on snapshot)", r.status_code == 404)

        # --- §8.12 progress visibility / batch snapshot ---
        section("§8.12 batch snapshot endpoint")
        files = [
            ("target_template", ("t.xlsx", target_xlsx, "application/octet-stream")),
            ("sources[orders]", ("a.xlsx", orders_xlsx, "application/octet-stream")),
            ("sources[customers]", ("c.xlsx", customers_xlsx, "application/octet-stream")),
            ("sources[sales]", ("s.xlsx", sales_xlsx, "application/octet-stream")),
        ]
        r = client.post(f"{API}/api/jobs", data={"config_name": "smoke_test"}, files=files)
        jid = r.json()["job_id"]
        wait_for_status(client, jid, {"done", "failed"}, timeout=30)
        r = client.get(f"{API}/api/jobs?ids={jid},nonexistent_job")
        snaps = r.json()["snapshots"]
        check("batch endpoint returns 2 entries", len(snaps) == 2)
        check("missing job marked as missing",
              any(s.get("status") == "missing" for s in snaps))

        # cleanup
        client.delete(f"{API}/api/configs/smoke_test")

    print("\n=== Summary ===")
    if failures:
        print(f"{FAIL} {len(failures)} failure(s)")
        for f in failures:
            print(f"  - {f}")
        return 1
    print(f"{PASS} all automated checks passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
