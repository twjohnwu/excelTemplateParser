"""Fetch real wholesale agricultural trade data from Taiwan MOA Open Data.

Source: https://data.moa.gov.tw/api/v1/AgriProductsTransType/
License: 政府資料開放授權條款 1.0 (compatible with CC BY 4.0)

Run from this directory with the backend venv active:
    cd examples/02_agri_market_report
    ../../backend/.venv/bin/python fetch.py

Produces:
    sources/daily_trades.xlsx     — 1 day of wholesale trades from real data
    sources/market_codes.xlsx     — manually curated: market code → city
    sources/tc_type_codes.xlsx    — manually curated: TcType → 分類
    template.xlsx                 — target template

The trade date is fixed (TARGET_DATE) so re-running gives the same output.
"""
from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
SOURCES = HERE / "sources"

# ROC 114.06.18 = 2025-06-18. Fixed so the example is reproducible.
TARGET_DATE = "114.06.18"
API_URL = (
    "https://data.moa.gov.tw/api/v1/AgriProductsTransType/"
    f"?Start_time={TARGET_DATE}&End_time={TARGET_DATE}&top=5000"
)

# Manually curated from public market info. The API returns market codes,
# but no companion lookup table, so this file plays the role of "我的對照表".
MARKETS = [
    ("104", "台北二", "台北市"),
    ("105", "台北市場", "台北市"),
    ("109", "台北一", "台北市"),
    ("220", "板橋區", "新北市"),
]

# TcType is the trade-category code on the API. The mapping comes from public
# wholesale-market documentation; see README for the source link.
TC_TYPES = [
    ("N04", "蔬菜"),
    ("N05", "水果"),
    ("N06", "花卉"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2D6A4F")
CENTER = Alignment(horizontal="center", vertical="center")


def _write(path: Path, sheet: str, headers: list[str], rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for r in rows:
        ws.append(r)
    for idx, h in enumerate(headers, start=1):
        col = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col].width = max(12, len(str(h)) * 1.8)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def fetch_trades() -> list[dict]:
    """Fetch trade rows from the MOA Open Data API.

    Uses curl rather than urllib because the MOA endpoint serves a cert
    whose chain trips Python's stricter SSL validation ("Missing Subject
    Key Identifier") while curl accepts it.
    """
    result = subprocess.run(
        ["curl", "-sS", "--max-time", "60", API_URL],
        capture_output=True,
        text=True,
        check=True,
    )
    payload = json.loads(result.stdout)
    data = payload.get("Data", [])
    if not data:
        raise RuntimeError(f"API returned no rows for {TARGET_DATE}")
    return data


def write_daily_trades(rows: list[dict]) -> None:
    # Use API field names verbatim — the example demonstrates that the tool can
    # consume third-party schemas without renaming the source.
    headers = [
        "TransDate",
        "TcType",
        "CropCode",
        "CropName",
        "MarketCode",
        "MarketName",
        "Upper_Price",
        "Middle_Price",
        "Lower_Price",
        "Avg_Price",
        "Trans_Quantity",
    ]
    data_rows = [tuple(r.get(h, "") for h in headers) for r in rows]
    _write(SOURCES / "daily_trades.xlsx", "trades", headers, data_rows)


def write_market_codes() -> None:
    _write(
        SOURCES / "market_codes.xlsx",
        "市場",
        ["市場代碼", "市場名稱", "所在縣市"],
        MARKETS,
    )


def write_tc_type_codes() -> None:
    _write(
        SOURCES / "tc_type_codes.xlsx",
        "TcType",
        ["TcType", "分類"],
        TC_TYPES,
    )


def write_template() -> None:
    _write(
        HERE / "template.xlsx",
        "市場日報",
        ["交易日", "所在縣市", "市場", "分類", "作物", "平均價", "交易量(公斤)"],
        [],
    )


def main() -> None:
    print(f"Fetching {API_URL} ...")
    rows = fetch_trades()
    print(f"  got {len(rows)} rows")
    write_daily_trades(rows)
    write_market_codes()
    write_tc_type_codes()
    write_template()
    print("Generated:")
    for p in sorted(HERE.rglob("*.xlsx")):
        print(f"  {p.relative_to(HERE)}")


if __name__ == "__main__":
    main()
