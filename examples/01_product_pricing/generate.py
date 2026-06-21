"""Generate mock data for the 01_product_pricing example.

Run from this directory with the backend venv active:
    cd examples/01_product_pricing
    ../../backend/.venv/bin/python generate.py

Produces:
    sources/product_master.xlsx      — 20 SKUs (primary)
    sources/supplier_A_quote.xlsx    — 12 SKUs (Chinese column names)
    sources/supplier_B_quote.xlsx    — 13 SKUs (English column names)
    sources/supplier_C_quote.xlsx    — 13 SKUs (mixed column names)
    template.xlsx                    — target template with styled header

Coverage is intentionally uneven so the outer-join output shows products
quoted by all three, by some, and by none.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
SOURCES = HERE / "sources"

# 20 SKUs, computer peripherals only (no brand names)
PRODUCTS = [
    ("SKU001", "USB Type-A 集線器 4 port",      "集線器",   599),
    ("SKU002", "USB Type-C 集線器 7-in-1",      "集線器",  1290),
    ("SKU003", "無線滑鼠 標準款",                "輸入設備",  690),
    ("SKU004", "無線滑鼠 人體工學版",            "輸入設備", 1190),
    ("SKU005", "機械鍵盤 紅軸",                  "輸入設備", 2490),
    ("SKU006", "機械鍵盤 茶軸",                  "輸入設備", 2490),
    ("SKU007", "HD 視訊鏡頭 1080p",              "影音設備", 1290),
    ("SKU008", "4K 視訊鏡頭",                    "影音設備", 2890),
    ("SKU009", "桌上型電容麥克風",                "影音設備", 1690),
    ("SKU010", "抗噪耳機罩",                      "影音設備",  990),
    ("SKU011", "隨身碟 64GB",                    "儲存裝置",  390),
    ("SKU012", "隨身碟 128GB",                   "儲存裝置",  690),
    ("SKU013", "行動硬碟 1TB",                   "儲存裝置", 1990),
    ("SKU014", "行動硬碟 2TB",                   "儲存裝置", 2990),
    ("SKU015", "SSD 外接盒 NVMe",                "儲存裝置",  890),
    ("SKU016", "HDMI 線 1.5m",                   "線材",     290),
    ("SKU017", "HDMI 線 3m",                     "線材",     490),
    ("SKU018", "DisplayPort 線 1.5m",            "線材",     590),
    ("SKU019", "USB-C to HDMI 轉接器",            "線材",     790),
    ("SKU020", "桌面集線管理盒",                  "集線器",   350),
]

# Coverage plan — different suppliers serve different SKU ranges
COVERAGE_A = range(1, 13)   # SKU001 – SKU012  (12 items, cheapest)
COVERAGE_B = range(6, 19)   # SKU006 – SKU018  (13 items, mid-tier)
COVERAGE_C = range(3, 16)   # SKU003 – SKU015  (13 items)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="3B5998")
CENTER = Alignment(horizontal="center", vertical="center")


def _write(path: Path, sheet_name: str, headers: list[str], rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for row in rows:
        ws.append(row)
    # Best-effort column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(12, len(str(header)) * 2)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_product_master() -> None:
    rows = [(sku, name, cat, price) for sku, name, cat, price in PRODUCTS]
    _write(
        SOURCES / "product_master.xlsx",
        sheet_name="商品主檔",
        headers=["SKU", "商品名稱", "分類", "建議售價"],
        rows=rows,
    )


def _supplier_price(sku_index: int, base: int, factor: float) -> int:
    # Deterministic, no Random — different multipliers per supplier for realism
    return int(round(base * factor)) + (sku_index % 5) * 5


def _supplier_stock(sku_index: int, base: int) -> int:
    # Deterministic stock variation
    return base + (sku_index * 7) % 60


def write_supplier_A() -> None:
    rows = []
    for i in COVERAGE_A:
        sku, name, _cat, price = PRODUCTS[i - 1]
        rows.append(
            (sku, name, _supplier_price(i, price, 0.65), _supplier_stock(i, 40))
        )
    _write(
        SOURCES / "supplier_A_quote.xlsx",
        sheet_name="月報價",
        headers=["貨號", "品名", "單價", "庫存"],
        rows=rows,
    )


def write_supplier_B() -> None:
    rows = []
    for i in COVERAGE_B:
        sku, name, _cat, price = PRODUCTS[i - 1]
        rows.append(
            (sku, name, _supplier_price(i, price, 0.70), _supplier_stock(i, 20))
        )
    _write(
        SOURCES / "supplier_B_quote.xlsx",
        sheet_name="Monthly Quote",
        headers=["SKU", "Product Name", "Price", "Stock"],
        rows=rows,
    )


def write_supplier_C() -> None:
    rows = []
    for i in COVERAGE_C:
        sku, name, _cat, price = PRODUCTS[i - 1]
        rows.append(
            (sku, name, _supplier_price(i, price, 0.68), _supplier_stock(i, 60))
        )
    _write(
        SOURCES / "supplier_C_quote.xlsx",
        sheet_name="報價單",
        headers=["商品編號", "商品名", "Unit Price", "可用庫存"],
        rows=rows,
    )


def write_template() -> None:
    _write(
        HERE / "template.xlsx",
        sheet_name="比價表",
        headers=["SKU", "商品名稱", "分類", "建議售價", "A報價", "B報價", "C報價"],
        rows=[],
    )


def main() -> None:
    write_product_master()
    write_supplier_A()
    write_supplier_B()
    write_supplier_C()
    write_template()
    print("Generated:")
    for p in sorted(HERE.rglob("*.xlsx")):
        print(f"  {p.relative_to(HERE)}")


if __name__ == "__main__":
    main()
